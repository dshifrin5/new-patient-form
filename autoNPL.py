import subprocess
import time
import requests
import re
import os
import base64
import traceback
from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime
from openpyxl import load_workbook, Workbook




# === INIT FLASK APP ===
app = Flask(__name__)
CORS(app, origins=[
    "https://new-patient-form.onrender.com", 
    "https://f096-47-18-38-63.ngrok-free.app"
], supports_credentials=True)


# === AUTO DEPLOY SECTION ===
RENDER_SERVICE_ID = "srv-d0r3fbfdiees73blq330"
RENDER_API_KEY = "rnd_PVTkldXhRNq8DwWypV9FzBihMVjd"


def start_ngrok():
    # Kill any existing ngrok instances
    subprocess.run(["taskkill", "/f", "/im", "ngrok.exe"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    ngrok_proc = subprocess.Popen(["ngrok", "http", "5002"], stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    time.sleep(5)
    response = requests.get("http://localhost:4040/api/tunnels")
    public_url = response.json()["tunnels"][0]["public_url"]
    print(f"[+] Ngrok URL: {public_url}")
    return ngrok_proc, public_url


def update_index_html(public_url):
    with open("index.html", "r", encoding="utf-8") as file:
        html = file.read()
    updated = re.sub(r"https?://[a-z0-9\-]+\.ngrok-free\.app|http://localhost:5002", public_url, html)
    if html != updated:
        with open("index.html", "w", encoding="utf-8") as file:
            file.write(updated)
        print("[+] index.html updated with new ngrok URL")
        return True
    else:
        print("[~] No changes needed — URL already current.")
        return False

def commit_and_push_changes():
    subprocess.run(["git", "add", "index.html"], check=True)
    subprocess.run(["git", "commit", "-m", "Auto-update ngrok URL"], check=False)
    subprocess.run(["git", "push", "--set-upstream", "origin", "main"], check=False)
    print("[+] Pushed to GitHub")

def trigger_render_deploy():
    headers = {
        "Authorization": f"Bearer {RENDER_API_KEY}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    response = requests.post(
        f"https://api.render.com/v1/services/{RENDER_SERVICE_ID}/deploys",
        headers=headers
    )
    if response.status_code == 201:
        print("[🚀] Render deployment triggered successfully.")
    else:
        print(f"[!] Failed to trigger deployment: {response.text}")

def auto_deploy():
    ngrok_proc, new_url = start_ngrok()
    changed = update_index_html(new_url)
    if changed:
        commit_and_push_changes()
    else:
        print("[~] Skipped commit and push.")
    trigger_render_deploy()

    print("[⚠️] Ngrok will now remain running. Do NOT terminate this script.")
    print("[🌐] Leave this terminal open to keep ngrok and your server alive.")
    print("[🔗] Public URL:", new_url)
    return ngrok_proc  # optionally return for manual control

fields_in_order = [
    "first_name", "last_name", "middle_initial", "gender", "family_status", "birth_date", "email",
    "mobile", "home_phone", "call_time", "address1", "address2", "city", "state", "zip",
    "insured_first", "insured_last", "insured_birth", "insured_employer", "plan_name",
    "relationship", "insurance_id", "insurance_card", "insurance_consent", "has_secondary",
    "secondary_card", "responsible_choice", "parent_first_name", "parent_last_name",
    "parent_gender", "parent_family_status", "parent_birth_date", "parent_email",
    "physician_name", "pharmacy_info", "serious_illness", "illness_description",
    "blood_transfusion", "transfusion_date", "pregnant", "nursing", "birth_control",
    "anemia", "arthritis", "artificial_joints", "asthma",
    "blood_disease", "cancer", "cerebral_palsy", "diflucan",
    "development_delay", "diabetes", "ehlers_danlos_syndrome",
    "epilepsy", "fainting", "genetic_syndrome", "glaucoma",
    "hivaids", "hashimoto_disease", "head_injuries",
    "heart_conditions", "heart_disease", "heart_murmur",
    "hemophilia", "hepatitis", "high_blood_pressure",
    "jaundice", "joint_conditions", "kidney_disease",
    "latent_tb", "liver_disease", "muscular_dystrophy",
    "nervous_disorders", "neutropenia", "pacemaker",
    "radiation_treatment", "respiratory_problems", "rheumatic_fever",
    "rheumatism", "scarlet_fever", "seizure",
    "shortness_of_breath", "sinus_problems", "situs_inversus",
    "spina_bifida", "stomach_problems", "stroke",
    "thyroid_problems", "tuberculosis", "tumors",
    "ulcers", "venereal_disease", "williams_syndrome",
    "taking_medication", "medications", "allergies", "allergy_details", "final_acknowledgement",
    # Consents
    "consentService", "consentCancel", "consentHIPAA", "financialResponsibility",
    "consentTreatmentRelease", "consentDental", "consentMedications", "consentPlanChange",
    "consentTeethRemoval", "consentSurgery",
    # Final Signature
    "formCompleterName", "relationshipToPatient",
]


# === FORM PROCESSING ===
SAVE_FOLDER = "submissions"
EXCEL_PATH = os.path.join(SAVE_FOLDER, "submissions.xlsx")
os.makedirs(SAVE_FOLDER, exist_ok=True)


if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(fields_in_order)
    wb.save(EXCEL_PATH)



from openpyxl import load_workbook, Workbook


# assume SAVE_FOLDER and EXCEL_PATH are defined earlier, e.g.:
# SAVE_FOLDER = "submissions"
# EXCEL_PATH = os.path.join(SAVE_FOLDER, "submissions.xlsx")
# and fields_in_order is your full list of column keys

@app.route("/submit", methods=["POST", "OPTIONS"])
def submit():
    if request.method == "OPTIONS":
        return jsonify({"status": "ok"}), 200

    try:
        # 1) Parse JSON
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data received"}), 400

        # 2) Fresh timestamp for PNG filename
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

        # 3) Optionally save signature PNG
        sig = data.get("signature", "")
        if isinstance(sig, str) and sig.startswith("data:image"):
            try:
                png_data = sig.split(",", 1)[1]
                path = os.path.join(SAVE_FOLDER, f"signature_{timestamp}.png")
                with open(path, "wb") as f:
                    f.write(base64.b64decode(png_data))
            except Exception:
                print("[!] Bad signature data; skipping PNG")
        else:
            print("[~] No signature provided; skipping PNG")

        # 4) Load the Excel workbook
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # 5) Normalize medical-history checkboxes
        condition_field_keys = [
            "anemia", "arthritis", "artificial_joints", "asthma",
            "blood_disease", "cancer", "cerebral_palsy", "diflucan",
            "development_delay", "diabetes", "ehlers_danlos_syndrome",
            "epilepsy", "fainting", "genetic_syndrome", "glaucoma",
            "hivaids", "hashimoto_disease", "head_injuries",
            "heart_conditions", "heart_disease", "heart_murmur",
            "hemophilia", "hepatitis", "high_blood_pressure",
            "jaundice", "joint_conditions", "kidney_disease",
            "latent_tb", "liver_disease", "muscular_dystrophy",
            "nervous_disorders", "neutropenia", "pacemaker",
            "radiation_treatment", "respiratory_problems", "rheumatic_fever",
            "rheumatism", "scarlet_fever", "seizure",
            "shortness_of_breath", "sinus_problems", "situs_inversus",
            "spina_bifida", "stomach_problems", "stroke",
            "thyroid_problems", "tuberculosis", "tumors",
            "ulcers", "venereal_disease", "williams_syndrome"
        ]
        binary_fields = ["pregnant", "nursing", "birth_control"] + condition_field_keys

        for field in binary_fields:
            val = str(data.get(field, "")).strip().lower()
            data[field] = "Yes" if val in ("yes", "true", "on") else ""

        for key in ("insurance_card", "secondary_card"):
                    if isinstance(data.get(key), dict):
                        data[key] = ""

        # 6) Build the row in the order of fields_in_order
        row = [data.get(col, "") for col in fields_in_order]
        ws.append(row)

        # 7) Save back to disk
        wb.save(EXCEL_PATH)

        return jsonify({"status": "success"}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500



# === START APP ===
if __name__ == "__main__":
    try:
        ngrok_proc = auto_deploy()
    except Exception as e:
        print("[Auto Deploy Failed]", e)

    print("[✔️] Starting Flask...")
    app.run(host="0.0.0.0", port=5002)
